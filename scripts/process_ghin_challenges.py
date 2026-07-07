#!/usr/bin/env python3
"""Build processed GHIN Challenges data from official GHIN Challenges workbooks."""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ACTIVE_WORKBOOK = Path("/Users/EvanBelfi/Downloads/GHIN_Challenges_ACTIVE.xlsx")
COMPLETED_WORKBOOK = Path("/Users/EvanBelfi/Downloads/GHIN_Challenges_COMPLETED.xlsx")
UPCOMING_WORKBOOK = Path("/Users/EvanBelfi/Downloads/GHIN_Challenges_UPCOMING.xlsx")
OUTPUT_PATH = PROJECT_ROOT / "data" / "processed" / "ghin_challenges.json"


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def as_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(value))


def as_iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value).date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def read_qa_summary(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    qa_sheet = next((name for name in wb.sheetnames if "QA" in name), None)
    if not qa_sheet:
        return {}
    rows = list(wb[qa_sheet].iter_rows(values_only=True))
    summary: dict[str, Any] = {}
    for label, value, *_ in rows[1:]:
        if label:
            summary[str(label)] = value
    return summary


def read_challenges(path: Path, sheet_name: str, status: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_header(cell) for cell in next(rows)]
    records: list[dict[str, Any]] = []
    for row in rows:
        values = dict(zip(headers, row))
        name = values.get("name") or values.get("challenge_name")
        aga = values.get("aga")
        if not name or not aga:
            continue
        golfers = as_int(values.get("golfers"))
        ranked_golfers = as_int(values.get("ranked_golfers"))
        scores_posted = as_int(values.get("scores_posted"))
        records.append(
            {
                "name": str(name).strip(),
                "aga": str(aga).strip(),
                "status": status,
                "startDate": as_iso_date(values.get("start_date")),
                "endDate": as_iso_date(values.get("end_date")),
                "golfers": golfers,
                "rankedGolfers": ranked_golfers,
                "scoresPosted": scores_posted,
                "rankedGolferRate": ranked_golfers / golfers if golfers else None,
                "scoresPerGolfer": scores_posted / golfers if golfers else None,
            }
        )
    return records


def workbook_sheet(path: Path, preferred: tuple[str, ...]) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in preferred:
        if sheet_name in wb.sheetnames:
            return sheet_name
    return next(name for name in wb.sheetnames if "QA" not in name)


def read_optional_challenges(path: Path, preferred_sheets: tuple[str, ...], status: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return read_challenges(path, workbook_sheet(path, preferred_sheets), status)


def read_upcoming_challenges(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    sheet_name = workbook_sheet(path, ("Upcoming Challenges", "GHIN Challenges"))
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_header(cell) for cell in next(rows)]
    upcoming_rows = []
    for row in rows:
        values = dict(zip(headers, row))
        scheduled_status = values.get("scheduled_status") or values.get("status")
        if str(scheduled_status or "").strip().lower() == "upcoming":
            upcoming_rows.append(row)
    return read_challenges_from_rows(headers, upcoming_rows, "Upcoming")


def read_challenges_from_rows(headers: list[str], rows: list[tuple[Any, ...]], status: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        values = dict(zip(headers, row))
        name = values.get("name") or values.get("challenge_name")
        aga = values.get("aga")
        if not name or not aga:
            continue
        golfers = as_int(values.get("golfers"))
        ranked_golfers = as_int(values.get("ranked_golfers"))
        scores_posted = as_int(values.get("scores_posted"))
        records.append(
            {
                "name": str(name).strip(),
                "aga": str(aga).strip(),
                "status": status,
                "startDate": as_iso_date(values.get("start_date")),
                "endDate": as_iso_date(values.get("end_date")),
                "golfers": golfers,
                "rankedGolfers": ranked_golfers,
                "scoresPosted": scores_posted,
                "rankedGolferRate": ranked_golfers / golfers if golfers else None,
                "scoresPerGolfer": scores_posted / golfers if golfers else None,
            }
        )
    return records


def metric_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    golfers = sum(row["golfers"] for row in records)
    ranked_golfers = sum(row["rankedGolfers"] for row in records)
    scores_posted = sum(row["scoresPosted"] for row in records)
    active_challenges = sum(1 for row in records if row["status"] == "Active")
    completed_challenges = sum(1 for row in records if row["status"] == "Completed")
    upcoming_challenges = sum(1 for row in records if row["status"] == "Upcoming")
    return {
        "activeChallenges": active_challenges,
        "completedChallenges": completed_challenges,
        "upcomingChallenges": upcoming_challenges,
        "totalChallenges": active_challenges + completed_challenges + upcoming_challenges,
        "totalGolfers": golfers,
        "rankedGolfers": ranked_golfers,
        "scoresPosted": scores_posted,
        "rankedGolferRate": ranked_golfers / golfers if golfers else None,
        "scoresPerGolfer": scores_posted / golfers if golfers else None,
    }


def status_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {"status": status, **metric_summary([row for row in records if row["status"] == status])}
        for status in ("Active", "Completed", "Upcoming")
    ]


def executive_summary(records: list[dict[str, Any]], upcoming_records: list[dict[str, Any]]) -> dict[str, Any]:
    summary = metric_summary(records)
    summary["upcomingChallenges"] = len(upcoming_records)
    summary["totalChallenges"] = (
        summary["activeChallenges"]
        + summary["completedChallenges"]
        + summary["upcomingChallenges"]
    )
    return summary


def top_agas(records: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        grouped[row["aga"]].append(row)
    rows = []
    for aga, aga_records in grouped.items():
        rows.append({"aga": aga, **metric_summary(aga_records)})
    return sorted(rows, key=lambda row: (-row["totalGolfers"], row["aga"]))[:limit]


def top_challenges(records: list[dict[str, Any]], metric: str, limit: int = 10) -> list[dict[str, Any]]:
    return sorted(records, key=lambda row: (-row[metric], row["name"]))[:limit]


def main() -> None:
    active_records = read_challenges(ACTIVE_WORKBOOK, "GHIN Challenges", "Active")
    completed_records = read_challenges(COMPLETED_WORKBOOK, "Completed Challenges", "Completed")
    upcoming_records = read_upcoming_challenges(UPCOMING_WORKBOOK)
    records = active_records + completed_records

    processed = {
        "metadata": {
            "schemaVersion": 1,
            "status": "official",
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sources": [
                {
                    "file": ACTIVE_WORKBOOK.name,
                    "status": "Active",
                    "qaSummary": read_qa_summary(ACTIVE_WORKBOOK),
                },
                {
                    "file": COMPLETED_WORKBOOK.name,
                    "status": "Completed",
                    "qaSummary": read_qa_summary(COMPLETED_WORKBOOK),
                },
                {
                    "file": UPCOMING_WORKBOOK.name,
                    "status": "Upcoming",
                    "found": UPCOMING_WORKBOOK.exists(),
                    "qaSummary": read_qa_summary(UPCOMING_WORKBOOK) if UPCOMING_WORKBOOK.exists() else {},
                },
            ],
        },
        "summary": executive_summary(records, upcoming_records),
        "statusSummary": status_summary(records + upcoming_records),
        "topAgasByGolfers": top_agas(records),
        "topChallengesByGolfers": top_challenges(records, "golfers"),
        "topChallengesByScoresPosted": top_challenges(records, "scoresPosted"),
        "challenges": records,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(processed, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(json.dumps(processed["summary"], indent=2))


if __name__ == "__main__":
    main()
